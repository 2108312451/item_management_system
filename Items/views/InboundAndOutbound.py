from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import datetime
from Items.models import Items,Warehousing,Outbound
from Function.models import Appointment,Collections
from Items.serializers.ItemsDataSerializers import ItemsSerializers,WarehousingSerializers,OutboundSerializers
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from PIL import Image
import io
import os
from django.http import FileResponse
from io import BytesIO
from openpyxl import load_workbook

# 单个
class ItemOperateView(APIView):
    # 单个添加物品/入库
    def post(self,request):
        # 获取图片数据
        image_data = request.data.get('image')

        # 将图片数据转换为文件对象
        image_file = ContentFile(image_data.read())

        # 保存图片文件到磁盘上的指定路径（比如 static/images 文件夹）
        file_path = default_storage.save('static/images/' + request.data.get('item_name') + '.jpg', image_file)

        item = Items.objects.create(item_name=request.data.get('item_name'),item_code=request.data.get('item_code'),
                                    value=request.data.get('value'),specifications=request.data.get('specifications'),
                                    category=request.data.get('category'),brand=request.data.get('brand'),
                                    inventory=request.data.get('inventory'),campus=request.data.get('campus'),
                                    location=request.data.get('location'),max_quantity=request.data.get('max_quantity'),
                                    instructions=request.data.get('instructions'),pictureUrl=image_file,#图片地址（单独传入图片）
                                    approval_classification=request.data.get('approval_classification'))
        item.save()
        # 获取当前时间
        current_time = datetime.datetime.now()
        # 将时间转换为字符串
        time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
        warehous = Warehousing.objects.create(item_name=request.data.get('item_name'),item_code=request.data.get('item_code'),
                                              category=request.data.get('category'),brand=request.data.get('brand'),
                                              numbers=request.data.get('inventory'),approval_classification=request.data.get('approval_classification'),
                                              times=time_str)

        warehous.save()
        if request.data.get('category') == '设备':
            # 获取当前日期
            current_date = datetime.datetime.now().date()
            # 创建未来5天预约时间段
            for i in range(1,6):
                # 将日期加一天
                next_day = current_date + datetime.timedelta(days=i)
                # 将日期转换为字符串
                date_str = next_day.strftime('%Y-%m-%d')
                obj = Appointment.objects.create(item_id=item.id,item_name=item.item_name,time=date_str,day=i)
                obj.save()

        return Response({"ok_create": True}, status=status.HTTP_200_OK)

    #修改物品信息
    def put(self,request):
        try:
            image_data = request.data.get('image')
            if image_data:
                # 定义路径和文件名
                folder_path = 'static/images/'
                file_name = request.data.get('item_name') + '.jpg'
                file_path = folder_path + file_name

                # 判断文件是否存在
                if os.path.exists(file_path):
                    # 存在同名文件，删除文件
                    os.remove(file_path)
                else:
                    pass
                # 将图片数据转换为文件对象
                image_file = ContentFile(image_data.read())
                # 保存图片文件到磁盘上的指定路径（比如 static/images 文件夹）
                default_storage.save('static/images/' + request.data.get('item_name') + '.jpg', image_file)
                obj = Items.objects.get(id=request.data.get("id"))
                obj.pictureUrl = image_file  # 图片地址（单独传入图片）
                obj.save()
            else:
                pass
        except:
            pass
        try:
            obj = Items.objects.get(id=request.data.get("id"))
            obj.item_name = request.data.get('item_name')
            obj.item_code = request.data.get('item_code')
            obj.value = request.data.get('value')
            obj.specifications = request.data.get('specifications')
            obj.category = request.data.get('category')
            obj.brand = request.data.get('brand')
            obj.inventory = request.data.get('inventory')
            obj.campus = request.data.get('campus')
            obj.location = request.data.get('location')
            obj.max_quantity = request.data.get('max_quantity')
            obj.instructions = request.data.get('instructions')
            obj.approval_classification = request.data.get('approval_classification')
            obj.save()
        except:
            pass
        return Response({"ok_put":True},status=status.HTTP_200_OK)

    def get(self,request,itemid,username):
        itemobj = Items.objects.get(id=itemid)
        itemdata = ItemsSerializers(instance=itemobj,many=False)
        exists = Collections.objects.filter(item_id=itemid, username=username).exists()
        colloid = 0
        if exists:
            colloid = Collections.objects.get(item_id=itemid, username=username).id
        return Response({"itemdata":itemdata.data,"collection":exists,"colloid":colloid})

# 获取表数据
class GetItemData(APIView):
    def get(self,request,pk):
        # 物品表
        if pk == 0:
            item = Items.objects.all()
            itemdata = ItemsSerializers(instance=item,many=True)
            return Response({"itemdata":itemdata.data}, status=status.HTTP_200_OK)
        # 入库表
        elif pk == 1:
            warehous = Warehousing.objects.all()
            warehousdata = WarehousingSerializers(instance=warehous,many=True)
            return Response({"itemdata": warehousdata.data}, status=status.HTTP_200_OK)
        #出库表
        elif pk == 2:
            outbound = Outbound.objects.all()
            outbounddata = OutboundSerializers(instance=outbound,many=True)
            return Response({"outbounddata": outbounddata.data}, status=status.HTTP_200_OK)
        else:
            return Response({"mesage": "pk值为空"}, status=status.HTTP_200_OK)

# 删除物品出库/报废
class DeleteItem(APIView):

    #删除物品/出库
    def delete(self,request,pk,id):
        # 删除物品中的一个（报废）
        if pk == 0:
            obj = Items.objects.get(id=id)
            if obj.inventory > 0:
                obj.inventory -= 1
                obj.save()

                # 获取当前时间
                current_time = datetime.datetime.now()
                # 将时间转换为字符串
                time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')

                outobj = Outbound.objects.create(item_name=obj.item_name,item_code=obj.item_code,
                                                 category=obj.category,brand=obj.brand,
                                                 numbers=1,approval_classification=obj.approval_classification,
                                                 Reason_Outbound=pk,times=time_str)
                outobj.save()
                return Response({"ok_del":True},status=status.HTTP_200_OK)
            else:
                return Response({"message":"库存不足"},status=status.HTTP_200_OK)
        # 直接删除物品
        elif pk == 1:
            obj = Items.objects.get(id=id)
            obj.delete()

            # 获取当前时间
            current_time = datetime.datetime.now()
            # 将时间转换为字符串
            time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')

            outobj = Outbound.objects.create(item_name=obj.item_name, item_code=obj.item_code,
                                             category=obj.category, brand=obj.brand,
                                             numbers=obj.inventory, approval_classification=obj.approval_classification,
                                             Reason_Outbound=pk, times=time_str)
            outobj.save()
            return Response({"ok_del": True}, status=status.HTTP_200_OK)

# 图片url
class ImageUrl(APIView):
    def get(self, request, imagename, format=None):
        try:
            image_path = f'static/images/{imagename}.jpg'  # 图片文件路径
            # 打开原始图片
            image = Image.open(image_path)
            width, height = image.size  # 获取原始图片的宽度和高度

            # 计算目标宽度和高度为原来的一半
            target_width = int(width * 1)
            target_height = int(height * 1)
            # 压缩图片
            image = image.resize((target_width, target_height))
            # 创建一个内存中的临时文件
            output = io.BytesIO()
            # 将压缩后的图片保存到临时文件
            image.save(output, format='JPEG')
            # 将临时文件对象移动到开头以便读取
            output.seek(0)
            # 返回压缩后的图片作为响应
            return FileResponse(output, content_type='image/jpeg')
        except IOError:
            return Response({'message':'未找到图片'},status=status.HTTP_404_NOT_FOUND)  # 如果未找到图片，则返回 404

# 批量创建物品/入库.
class MoreAdd(APIView):
    def post(self,request):
        try:
            excel_file = request.FILES['file']
            wb = load_workbook(excel_file)
            sheet = wb.active

            # 设置一个标志来跟踪当前行数
            first_row = 0

            for row in sheet.iter_rows(values_only=True):
                if first_row<1:
                    first_row += 1
                    continue  # 跳过第一行

                image_path = row[0]  # 图片文件路径

                # 读取图片文件的二进制数据
                with open(image_path, 'rb') as f:
                    image_data = f.read()

                # 将图片数据转换为文件对象
                image_file = ContentFile(image_data)

                # 保存图片文件到指定路径
                file_path = 'static/images/' + row[1] + '.jpg'  # 保存路径
                default_storage.save(file_path, image_file)

                # 逐行读取数据并存储到数据库
                data = {
                    'item_name': row[1],
                    'item_code': row[2],
                    'value': row[3],
                    'specifications': row[4],
                    'category': row[5],
                    'brand': row[6],
                    'inventory': row[7],
                    'campus': row[8],
                    'location': row[9],
                    'max_quantity': row[10],
                    'instructions': row[11],
                    'approval_classification': row[12],
                }
                objs = Items.objects.create(**data)
                objs.save()

            #入库
                # 获取当前时间
                current_time = datetime.datetime.now()
                # 将时间转换为字符串
                time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
                warehous = Warehousing.objects.create(item_name=row[1],
                                                      item_code=row[2],
                                                      category=row[5],
                                                      brand=row[6],
                                                      numbers=row[7],
                                                      approval_classification=row[12],
                                                      times=time_str)

                warehous.save()
                if row[5] == '设备':
                    # 获取当前日期
                    current_date = datetime.datetime.now().date()
                    # 创建未来5天预约时间段
                    for i in range(1, 6):
                        # 将日期加一天
                        next_day = current_date + datetime.timedelta(days=i)
                        # 将日期转换为字符串
                        date_str = next_day.strftime('%Y-%m-%d')
                        obj = Appointment.objects.create(item_id=objs.id, item_name=objs.item_name, time=date_str,
                                                         day=i)
                        obj.save()

            return Response({'ok_create': True},status=status.HTTP_200_OK)
        except:
            return Response({'message': False}, status=status.HTTP_400_BAD_REQUEST)
